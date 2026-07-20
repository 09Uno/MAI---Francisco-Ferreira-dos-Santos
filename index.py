# -*- coding: utf-8 -*-
"""
Conciliador bancário — Escritório F. Ferreira Advogados / Advbox (sem API)
=========================================================================
Fluxo (100% baseado em arquivo, sem integração via API):

    OFX do banco  +  Export do Financeiro do Advbox (.xlsx)
                       │
                       ▼
        cruza extrato x lançamentos do sistema
                       │
                       ▼
        gera  conciliacao.xlsx  com 3 abas:
          • DAR BAIXA        -> movimento já existe no Advbox (marcar pagamento)
          • CRIAR LANÇAMENTO -> não existe: já classificado e desdobrado (30%/repasse),
                                nas MESMAS colunas do export -> digitar/importar no Advbox
          • REVISAR          -> casos ambíguos / sem regra

Como usar:
    ext  = carregar_ofx("extrato_bb_abril.ofx", "I - BANCO DO BRASIL - FRANCISCO")
    ext += carregar_ofx("extrato_asaas.ofx",    "I - ASAAS - DR. FERREIRA")
    sistema = carregar_advbox_export("Advbox-MOVIMENTACAO-ABRIL.xlsx")
    gerar_planilha(ext, sistema, "conciliacao_abril.xlsx")
"""

import re
import unicodedata
from dataclasses import dataclass

# Patch openpyxl para tolerar xlsx do Advbox com margens inválidas (string vazia ao invés de float)
import openpyxl.descriptors.base as _desc
_orig_convert = _desc._convert
def _safe_convert(expected_type, value):
    if expected_type is float and (value is None or (isinstance(value, str) and value.strip() == "")):
        return 0.0
    return _orig_convert(expected_type, value)
_desc._convert = _safe_convert
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ------------------------------------------------------------------ CONFIG ----
PERC_HONORARIOS_PADRAO = 0.30      # 30% — ajustável por processo
JANELA_DIAS = 5                    # tolerância entre data do extrato e do lançamento

# ---------------------------------------------------------------- RULEBOOK ----
def norm(s) -> str:
    s = str(s or "").upper()
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).strip()

_RULES = [
    (r"TRANSFERENCIA (PARA|DE|ENTRE)", None, "TRANSFERÊNCIA ENTRE CONTAS"),
    (r"TAXA .*BOLETO|BOLETO COMPENSADO|TAXA .*MENSAG|TAXA .*WHATSAPP|TAXA .*SMS|"
     r"TAXA .*NOTIFICA|NOTIFICACAO (WHATSAPP|SMS)|TAXA .*MAQUININHA|TAXA .*PIX|"
     r"TAXA .*TRANSFER|TARIFA",
     None, "7. TAXAS BANCÁRIAS"),
    (r"JUROS (LIMITE|BANCARIO)", None, "7. JUROS BANCÁRIOS"),
    (r"TED ?INTERNET|CESTA DE SERVICO|PACOTE DE SERVICO", None, "7. TAXAS BANCÁRIAS"),
    (r"\bIOF\b", "DESPESA", "4. IOF"),
    (r"\bISS\b", "DESPESA", "4. ISS"),
    (r"\bINSS\b|DCTFWEB", "DESPESA", "4. INSS"),
    (r"SIMPLES NACIONAL", "DESPESA", "4. SIMPLES NACIONAL"),
    (r"\bFGTS\b", "DESPESA", "3. FGTS"),
    (r"VALE TRANSPORTE|\bVT\b", "DESPESA", "2. VALE TRANSPORTE"),
    (r"VALE ALIMENTAC|VALE ALIMETAC", "DESPESA", "2. VALE ALIMENTAÇÃO"),
    (r"ADIANTAMENTO", "DESPESA", "2. ADIANTAMENTOS"),
    (r"BONIFICAC", "DESPESA", "2. BONIFICAÇÃO"),
    (r"PRO ?-?LABORE|PROLABORE", "DESPESA", "2. PRÓ-LABORE"),
    (r"ACORDO DE SOCIO|RETIRADA DE SOCIO", "DESPESA", "2. ACORDO DE SÓCIOS"),
    (r"BENEFICIOS|FLASH", "DESPESA", "2. BENEFÍCIOS"),
    (r"SEGURO DE VIDA", "DESPESA", "2. SEGURO DE VIDA"),
    (r"SAUDE OCUPACIONAL|CONVENIO|EXAME ADMISSIONAL", "DESPESA", "2. SAÚDE OCUPACIONAL"),
    (r"RESCISAO|TRCT|ENCERRAMENTO DE ESTAGIO", "DESPESA", "2. RESCISÃO DO CONTRATO DE TRABALHO"),
    (r"REMUNERACAO DE ADVOGADOS", "DESPESA", "2. REMUNERAÇÃO DE ADVOGADOS PRESTADORES DE SERVIÇO"),
    (r"REMUNERACAO .*PARCEIRO|PARCERIA .*CLIENTE", "DESPESA", "2. REMUNERAÇÃO DE PARCEIROS - PRESTADORES DE SERVIÇO"),
    (r"SALARIO|BOLSA ESTAGIO|PRESTACAO DE SERVICOS", "DESPESA", "2. SALÁRIOS"),
    (r"ALUGUEL", "DESPESA", "5. ALUGUEL"),
    (r"ENEL|ENERGIA", "DESPESA", "5. ENERGIA ELÉTRICA"),
    (r"SABESP|CONTA DE AGUA|\bAGUA\b", "DESPESA", "5. ÁGUA"),
    (r"INTERNET|VIVO|CLARO|TELEFONIA|DIRECT CALL|FIBRA", "DESPESA", "5. INTERNET"),
    (r"SOFTWARE|3CX|CHAT GURU|GOOGLE DRIVE|CREDLOCALIZA", "DESPESA", "5. SOFTWARE E SISTEMAS"),
    (r"CONTABILIDADE|CONTADOR", "DESPESA", "5. CONTADOR"),
    (r"PAGAMENTO DE FATURA|PAGTO CARTAO|CARTAO DE CREDITO|CARTAO CREDITO", "DESPESA", "5. PAGAMENTO DE FATURA"),
    (r"SEGURO DO CARRO", "DESPESA", "5. SEGURO DO CARRO"),
    (r"ANUIDADE OAB", "DESPESA", "5. ANUIDADE OAB"),
    (r"ESTACIONAMENTO", "DESPESA", "5. ESTACIONAMENTO"),
    (r"COMODATO|PURIFICADOR|MAQUININHA REDE", "DESPESA", "5. SERVIÇOS DE COMODATO"),
    (r"MANUTENCAO PREVENTIVA|CONTRATO DE MANUTENCAO", "DESPESA", "5. INFORMÁTICA"),
    (r"DILIGENCIA|AUDIENCIA", "DESPESA", "6. DILIGÊNCIAS"),
    (r"CUSTAS|GUIA|TELEGRAMA|CARTA AR|CARTORIO", "DESPESA", "6. GUIA DE CUSTAS PAGAS PARA O CLIENTE"),
    (r"CONFRATERNIZAC|BOLO|ANIVERSARIO DO ESCRITORIO", "DESPESA", "6. CONFRATERNIZAÇÕES DO ESCRITÓRIO"),
    (r"CURSO|TREINAMENTO|MENTORIA|ADVOGADO 10X", "DESPESA", "6. CURSOS E TREINAMENTOS"),
    (r"INVESTIMENTO|CONSORCIO|PORTE VILELA|SALA TATUAPE", "DESPESA", "6. INVESTIMENTO - SALA TATUAPÉ"),
    (r"DIVISAO DE LUCROS", "DESPESA", "6. DIVISÃO DE LUCROS - MENSAL"),
    (r"MATERIAIS|COPOS", "DESPESA", "6. MATERIAIS DE LIMPEZA E ALIMENTAÇÃO"),
    (r"PASSAGEM|RODOVIARIA", "DESPESA", "6. PASSAGENS AÉREAS E RODOVIÁRIAS"),
    (r"MANUTENCAO|AR COMPRIME|HERMINIO", "DESPESA", "6. MANUTENÇÃO"),
    (r"TINTAS|OBRA|TECLADO|CAPINHA|COMPRA COM CARTAO|BALAO|GASTOS COM O ESCRITORIO",
     "DESPESA", "6. GASTOS COM O ESCRITÓRIO"),
    (r"REPASSE|#REPASSE", "DESPESA", "0. REPASSE DE CLIENTE"),
    (r"HONORARIOS DE PARCEIRO|GRATIFICACAO INDICACAO|PARCERIA .*DRA|PARCERIA .*DR ",
     "DESPESA", "0. HONORÁRIOS DE PARCEIROS"),
    (r"DEVOLUCAO", "DESPESA", "0. DEVOLUÇÃO DE HONORÁRIOS"),
    (r"REPASSE|#REPASSE", "RECEITA", "1. REPASSE DE CLIENTE"),
    (r"\bRPV\b|\bRPVS\b", "RECEITA", "1. RPVS"),
    (r"PRECATORIO", "RECEITA", "1. PRECATÓRIOS"),
    (r"LEV JUD|LEVANTAMENTO JUDICIAL|\bCR LEV\b|CREDITO JUDICIAL", "RECEITA", "1. ALVARÁS"),
    (r"(30%|35%|25%|20%).*ALVARA|HONORARIOS DO ALVARA|DO ALVARA", "RECEITA", "1. HONORÁRIOS FINAIS"),
    (r"ALVARA", "RECEITA", "1. ALVARÁS"),
    (r"SUCUMBENCIA", "RECEITA", "1. HONORÁRIOS DE SUCUMBÊNCIA"),
    (r"MULTA E JUROS|MULTAS E JUROS", "RECEITA", "1. MULTA E JUROS"),
    (r"CONSULTA|CONSULTORIA", "RECEITA", "1. HONORÁRIOS CONSULTORIAS"),
    (r"IMPOSTO DE RENDA|DECLARACAO IR", "RECEITA", "1. HONORÁRIOS - IMPOSTO DE RENDA"),
    (r"RENDIMENTO|APLICACAO|REEMBOLSO TAXA", "RECEITA", "1. APLICAÇÕES FINANCEIRAS"),
    (r"REEMBOLSO CUSTAS|REEMBOLSO CALCULO|REEMBOLSO CONVENIO|REEMBOLSO CUSTO",
     "RECEITA", "1. REEMBOLSO DE CUSTO POR CLIENTES"),
    (r"AJUSTE DE CONTA", "RECEITA", "1. AJUSTE DE CONTA"),
    (r"NAO IDENTIFICAD", "RECEITA", "1. PAGAMENTO  NÃO IDENTIFICADO"),
    (r"HONORARIOS INICIAIS|CONTRATUAIS INICIAIS|HONORARIOS CONTRATUAIS|"
     r"HONORARIOS PRESTACAO|DEFESA ADMINISTRATIVA|PRORROGACAO",
     "RECEITA", "1. HONORÁRIOS INICIAIS"),
    (r"HONORARIOS|ACORDO|TERMO DE ACORDO|30%|35%|25%|FGTS", "RECEITA", "1. HONORÁRIOS FINAIS"),
]
_RULES_C = [(re.compile(p), t, c) for p, t, c in _RULES]
_DESDOBRA = {"1. ALVARÁS", "1. RPVS", "1. PRECATÓRIOS"}
_AMBIGUO = re.compile(r"NEGOCIAD|NEGOCIACAO|PARCELA \d")


def classificar(descricao: str, tipo: str) -> Optional[str]:
    d = norm(descricao)
    for rx, t, cat in _RULES_C:
        if rx.search(d) and (t is None or t == tipo):
            return cat
    return None


# --------------------------------------------------------------- MODELOS ------
@dataclass
class MovExtrato:
    data: datetime
    valor: float          # + crédito (receita) / - débito (despesa)
    descricao: str
    conta: str
    id_banco: str = ""
    @property
    def tipo(self) -> str:
        return "RECEITA" if self.valor >= 0 else "DESPESA"


@dataclass
class Lancamento:
    conta: str; tipo: str; categoria: str; descricao: str; valor: float
    vencimento: datetime; pagamento: Optional[datetime] = None
    pessoa: str = ""; processo: str = ""; registro_interno: bool = False
    origem: str = ""; revisar: str = ""
    centro_custo: str = ""; setor: str = ""


# -------------------------------------------------------------- DESDOBRAMENTO --
def desdobrar(mov: MovExtrato, categoria: str, perc: float = PERC_HONORARIOS_PADRAO):
    V = abs(mov.valor)
    hon = round(V * perc, 2)
    repasse = round(V - hon, 2)
    cat_hon = "1. HONORÁRIOS FINAIS" if categoria == "1. ALVARÁS" else categoria
    _o = f"desdobramento:{mov.descricao[:40]}"
    return [
        Lancamento(conta=mov.conta, tipo="RECEITA", categoria=categoria, valor=V,
                   descricao="Valor identificado", vencimento=mov.data, pagamento=mov.data,
                   registro_interno=True, revisar="confirmar % de honorários e custas", origem=_o),
        Lancamento(conta=mov.conta, tipo="RECEITA", categoria=cat_hon, valor=hon,
                   descricao=f"{int(perc*100)}% honorários", vencimento=mov.data,
                   pagamento=mov.data, origem=_o),
        Lancamento(conta=mov.conta, tipo="DESPESA", categoria="0. REPASSE DE CLIENTE", valor=repasse,
                   descricao="Repasse crédito de cliente", vencimento=mov.data,
                   pagamento=mov.data, registro_interno=True, origem=_o),
    ]


# ------------------------------------------------------------- CARREGADORES ---
def _linha_de_saldo(memo: str) -> bool:
    m = norm(memo)
    return m.startswith("SALDO ") or "SALDO ANTERIOR" in m or "SALDO DIA" in m or \
        "SALDO FINAL" in m or "SALDO EM" in m or m in ("SALDO", "S A L D O")


def carregar_ofx(caminho: str, conta: str) -> list[MovExtrato]:
    """Parser OFX mínimo, sem dependências (SGML e XML)."""
    with open(caminho, "r", encoding="latin-1", errors="ignore") as f:
        txt = f.read()
    movs = []
    for bloco in re.findall(r"<STMTTRN>(.*?)</STMTTRN>", txt, re.S):
        def tag(t):
            m = re.search(rf"<{t}>([^<\r\n]+)", bloco)
            return m.group(1).strip() if m else ""
        memo = (tag("MEMO") or tag("NAME"))
        valor = float(tag("TRNAMT").replace(",", ".") or 0)
        if valor == 0 or _linha_de_saldo(memo):     # ignora saldos e linhas nulas
            continue
        dt = tag("DTPOSTED")[:8]
        movs.append(MovExtrato(
            data=datetime.strptime(dt, "%Y%m%d") if len(dt) == 8 else datetime.now(),
            valor=valor, descricao=memo, conta=conta, id_banco=tag("FITID")))
    return movs


def carregar_caixa_pdf(caminho: str, conta: str) -> list[MovExtrato]:
    """
    Lê o extrato da Caixa em PDF (a Caixa não fornece OFX).
    Formato: Data | Nr.Doc | Histórico | Valor D/C | Saldo D/C.
    Usa pdfplumber (pip) — sem depender do poppler.
    """
    import pdfplumber
    movs = []
    with pdfplumber.open(caminho) as pdf:
        for page in pdf.pages:
            txt = page.extract_text(layout=True) or ""
            for ln in txt.splitlines():
                m = re.match(r"\s*(\d{2}/\d{2}/\d{4})\s+(\d+)\s+(.+?)\s+"
                             r"([\d.]+,\d{2})\s+([DC])\s+[\d.]+,\d{2}\s+[DC]", ln)
                if not m:
                    continue
                data, _doc, hist, val, dc = m.groups()
                if _linha_de_saldo(hist):
                    continue
                v = float(val.replace(".", "").replace(",", "."))
                v = -v if dc == "D" else v
                movs.append(MovExtrato(data=datetime.strptime(data, "%d/%m/%Y"),
                                       valor=v, descricao=hist.strip(), conta=conta))
    return movs


def _patch_openpyxl():
    """O Advbox gera xlsx com margens de página vazias que quebram o openpyxl."""
    from openpyxl.descriptors import base as desc_base
    if getattr(desc_base, '_patched', False):
        return
    _orig = desc_base._convert
    def _safe(expected_type, value):
        if expected_type is float and (value is None or (isinstance(value, str) and value.strip() == '')):
            return 0.0
        return _orig(expected_type, value)
    desc_base._convert = _safe
    desc_base._patched = True


def _col(df, *nomes):
    """Busca coluna por nome, tolerando variações de acento/encoding."""
    for col in df.columns:
        c = norm(str(col))
        for n in nomes:
            if norm(n) == c:
                return col
    return nomes[0]


def carregar_advbox_export(caminho: str) -> list[dict]:
    """Lê o Excel exportado do Financeiro do Advbox e devolve os lançamentos."""
    _patch_openpyxl()
    raw = pd.read_excel(caminho, sheet_name=0, header=None, engine="openpyxl")
    hdr = None
    for i, row in raw.iterrows():
        vals = [norm(str(c)) for c in row.tolist()]
        if "CATEGORIA" in vals and ("DESCRICAO" in vals or "TIPO" in vals):
            hdr = i; break
    if hdr is None:
        raise ValueError("Cabeçalho não encontrado no export do Advbox.")
    df = raw.iloc[hdr + 1:].copy()
    df.columns = raw.iloc[hdr].tolist()
    col_tipo = _col(df, "Tipo")
    col_desc = _col(df, "Descrição", "Descricao")
    col_cat = _col(df, "Categoria")
    col_partes = _col(df, "Partes")
    col_pgto = _col(df, "Pagamento")
    col_vr = _col(df, "Valor recebido")
    col_vp = _col(df, "Valor pago")
    col_conta = _col(df, "Conta/Cartão", "Conta/Cartao", "Conta")
    col_cc = _col(df, "Centro de custo", "Centro de Custo")
    col_setor = _col(df, "Setor/Unidade", "Setor")
    df = df[df[col_tipo].isin(["RECEITA", "DESPESA"])]
    out = []
    for _, r in df.iterrows():
        val = r.get(col_vr) if r[col_tipo] == "RECEITA" else r.get(col_vp)
        try:
            val = float(val)
        except (TypeError, ValueError):
            continue
        pg = r.get(col_pgto)
        conta_cartao = str(r.get(col_conta, "") or "").strip()
        centro = str(r.get(col_cc, "") or "").strip()
        setor_un = str(r.get(col_setor, "") or "").strip()
        out.append({"valor": val,
                    "data": pd.to_datetime(pg, dayfirst=True, errors="coerce"),
                    "descricao": r.get(col_desc, ""),
                    "pessoa": r.get(col_partes, ""),
                    "categoria": r.get(col_cat, ""),
                    "conta": conta_cartao,
                    "centro_custo": centro,
                    "setor": setor_un})
    return out


# --------------------------------------------------------------- MATCHING -----
def _fuzzy(a: str, b: str) -> float:
    A, B = set(norm(a).split()), set(norm(b).split())
    return len(A & B) / max(1, len(A | B))


def conciliar(extrato: list[MovExtrato], sistema: list[dict]):
    baixas, novos, revisao = [], [], []
    usados = set()
    for mov in extrato:
        melhor, score = None, -1.0
        for i, l in enumerate(sistema):
            if i in usados or abs(abs(l["valor"]) - abs(mov.valor)) > 0.01:
                continue
            dt = l["data"]
            if pd.isna(dt) or abs((dt.to_pydatetime() - mov.data).days) > JANELA_DIAS:
                continue
            s = _fuzzy(l.get("descricao", ""), mov.descricao) + _fuzzy(l.get("pessoa", ""), mov.descricao)
            if s > score:
                melhor, score = i, s
        if melhor is not None:
            usados.add(melhor)
            baixas.append((mov, sistema[melhor]))
            continue
        cat = classificar(mov.descricao, mov.tipo)
        if cat is None or _AMBIGUO.search(norm(mov.descricao)):
            revisao.append((mov, cat))
            continue
        if cat in _DESDOBRA:
            novos.extend(desdobrar(mov, cat))
        else:
            novos.append(Lancamento(
                conta=mov.conta, tipo=mov.tipo, categoria=cat, descricao=mov.descricao,
                valor=abs(mov.valor), vencimento=mov.data, pagamento=mov.data,
                registro_interno=(cat == "0. REPASSE DE CLIENTE"),
                origem=f"extrato:{mov.id_banco or mov.descricao[:30]}"))
    return baixas, novos, revisao


# --------------------------------------------------------------- PLANILHA ------
_AZUL = "1F4E78"; _CINZA = "D9E1F2"; _AMAR = "FFF2CC"
_HEAD = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_TXT = Font(name="Arial", size=10)
_BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)
_D = "%d/%m/%Y"


def _estilo_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c); cell.font = _HEAD
        cell.fill = PatternFill("solid", fgColor=_AZUL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def gerar_planilha(extrato, sistema, saida="conciliacao.xlsx"):
    baixas, novos, revisao = conciliar(extrato, sistema)
    wb = Workbook()

    # ---- Resumo ----
    ws = wb.active; ws.title = "Resumo"
    ws.sheet_properties.tabColor = _AZUL
    ws["A1"] = "CONCILIAÇÃO BANCÁRIA"; ws["A1"].font = Font(name="Arial", bold=True, size=14)
    ws["A3"] = "Gerado em"; ws["B3"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    linhas = [("Movimentos no extrato", len(extrato)),
              ("Já no Advbox (dar baixa)", len(baixas)),
              ("Criar lançamento", len(novos)),
              ("Revisar", len(revisao))]
    for i, (k, v) in enumerate(linhas, start=5):
        ws.cell(i, 1, k).font = _TXT
        ws.cell(i, 2, v).font = Font(name="Arial", bold=True, size=10)
    ws["A11"] = "Total a criar (R$)"; ws["A11"].font = _TXT
    _nnovos = len(novos)
    ws["B11"] = (f"=SUM('CRIAR LANÇAMENTO'!E2:E{_nnovos + 1})" if _nnovos else 0)
    ws["B11"].number_format = 'R$ #,##0.00'
    for col, w in {"A": 30, "B": 22}.items():
        ws.column_dimensions[col].width = w

    # ---- DAR BAIXA ----
    ws = wb.create_sheet("DAR BAIXA"); ws.sheet_properties.tabColor = "548235"
    cols = ["Data extrato", "Valor", "Descrição (extrato)", "Conta",
            "Lançamento no Advbox", "Cliente", "Ação"]
    ws.append(cols)
    for mov, l in baixas:
        ws.append([mov.data.strftime(_D), abs(mov.valor), mov.descricao, mov.conta,
                   l.get("descricao", ""), l.get("pessoa", ""),
                   "Marcar pagamento com a data do extrato"])
    _estilo_header(ws, len(cols))

    # ---- CRIAR LANÇAMENTO (mesmas colunas do export do Advbox) ----
    ws = wb.create_sheet("CRIAR LANÇAMENTO"); ws.sheet_properties.tabColor = _AZUL
    cols = ["Conta/Cartão", "Tipo", "Categoria", "Descrição", "Valor",
            "Vencimento", "Pagamento", "Pessoa", "Processo", "Registro interno", "Revisar"]
    ws.append(cols)
    for l in novos:
        ws.append([l.conta, l.tipo, l.categoria, l.descricao, round(l.valor, 2),
                   l.vencimento.strftime(_D), l.pagamento.strftime(_D) if l.pagamento else "",
                   l.pessoa, l.processo, "Sim" if l.registro_interno else "",
                   l.revisar])
    _estilo_header(ws, len(cols))
    nlin = len(novos)
    # destaca linhas que precisam de revisão
    for i in range(2, nlin + 2):
        if ws.cell(i, 11).value:
            for c in range(1, len(cols) + 1):
                ws.cell(i, c).fill = PatternFill("solid", fgColor=_AMAR)

    # ---- REVISAR ----
    ws = wb.create_sheet("REVISAR"); ws.sheet_properties.tabColor = "C00000"
    cols = ["Data", "Valor", "Descrição", "Conta", "Tipo", "Sugestão", "Motivo"]
    ws.append(cols)
    for mov, cat in revisao:
        ws.append([mov.data.strftime(_D), abs(mov.valor), mov.descricao, mov.conta,
                   mov.tipo, cat or "—",
                   "sem regra" if cat is None else "ambíguo (inicial x final?)"])
    _estilo_header(ws, len(cols))

    # estilo geral de corpo + larguras
    for sh in wb.worksheets:
        if sh.title == "Resumo":
            continue
        for row in sh.iter_rows(min_row=2):
            for cell in row:
                cell.font = _TXT; cell.border = _BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=False)
        widths = {"A": 30, "B": 12, "C": 32, "D": 30, "E": 30, "F": 14,
                  "G": 12, "H": 26, "I": 20, "J": 14, "K": 30}
        for col, w in widths.items():
            sh.column_dimensions[col].width = w
        # formato de moeda na coluna de valor
        valcol = "E" if sh.title == "CRIAR LANÇAMENTO" else "B"
        for cell in sh[valcol][1:]:
            cell.number_format = 'R$ #,##0.00'

    wb.save(saida)
    return {"baixas": len(baixas), "novos": len(novos), "revisar": len(revisao)}


def executar_pasta(pasta: str = "."):
    """
    Modo 'clica e roda': lê TODOS os .ofx da pasta (o nome do arquivo vira o nome
    da conta) e o export do Advbox (o .xlsx que não é saída), e gera a planilha
    de conciliação com a data de hoje. Abre o resultado automaticamente no Windows.
    """
    import glob
    import os

    ofx_files = glob.glob(os.path.join(pasta, "*.ofx"))
    if not ofx_files and not glob.glob(os.path.join(pasta, "*.pdf")):
        print("⚠ Nenhum extrato (.ofx ou .pdf) encontrado nesta pasta.")
        print("  Coloque aqui o(s) extrato(s) do banco (ex.: 'BB Francisco.ofx').")
        return

    extrato = []
    for f in ofx_files:
        conta = os.path.splitext(os.path.basename(f))[0]   # nome do arquivo = conta
        movs = carregar_ofx(f, conta)
        extrato += movs
        print(f"  • Extrato OFX lido: {conta}  ({len(movs)} movimentos)")

    # extratos da Caixa vêm em PDF (a Caixa não fornece OFX)
    for f in glob.glob(os.path.join(pasta, "*.pdf")):
        conta = os.path.splitext(os.path.basename(f))[0]
        try:
            movs = carregar_caixa_pdf(f, conta)
            if movs:
                extrato += movs
                print(f"  • Extrato PDF (Caixa) lido: {conta}  ({len(movs)} movimentos)")
        except Exception as e:
            print(f"  ! Não consegui ler o PDF {os.path.basename(f)}: {e}")

    export = None
    for f in glob.glob(os.path.join(pasta, "*.xlsx")):
        if os.path.basename(f).lower().startswith("conciliacao"):
            continue
        try:
            sistema = carregar_advbox_export(f)
            export = f
            print(f"  • Export do Advbox lido: {os.path.basename(f)}  ({len(sistema)} lançamentos)")
            break
        except Exception:
            continue
    if export is None:
        print("⚠ Não encontrei o export do Advbox (.xlsx) nesta pasta.")
        print("  Exporte o Financeiro do Advbox para Excel e coloque o arquivo aqui.")
        return

    saida = os.path.join(pasta, "conciliacao_" + datetime.now().strftime("%Y-%m-%d") + ".xlsx")
    res = gerar_planilha(extrato, sistema, saida)
    print(f"\n✔ Pronto! Gerado: {os.path.basename(saida)}")
    print(f"    Dar baixa: {res['baixas']}  |  Criar: {res['novos']}  |  Revisar: {res['revisar']}")
    try:
        os.startfile(saida)          # abre no Excel (Windows)
    except Exception:
        pass


if __name__ == "__main__":
    import sys
    pasta = sys.argv[1] if len(sys.argv) > 1 else "."
    print("=== Conciliação bancária ===\n")
    executar_pasta(pasta)
    input("\nPressione ENTER para fechar.")