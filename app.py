# -*- coding: utf-8 -*-
"""
Interface web para o Conciliador Bancário — Advbox
"""

import json
import os
import uuid
from datetime import datetime

from dotenv import load_dotenv
from flask import (Flask, render_template, request, redirect, url_for,
                   send_file, flash, jsonify)

# importa o motor de conciliação e o client da API
import index as engine
from advbox_client import AdvboxClient

load_dotenv()  # carrega .env (ADVBOX_TOKEN)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", os.urandom(24))

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB

# Lista de categorias disponíveis (extraídas das regras)
CATEGORIAS = sorted({cat for _, _, cat in engine._RULES})

CENTROS_CUSTO = [
    "ADMINISTRATIVO", "CESSÃO DE CRÉDITO", "CIVEL",
    "PREVIDENCIARIO", "REPASSE", "TRABALHISTA",
]

SETORES = [
    "ADMINISTRATIVO", "ADMINISTRATIVO - TATUAPÉ", "CESSÃO DE CRÉDITO",
    "CIVEL", "COMERCIAL", "CONTROLADORIA JURÍDICA", "DIRETORIA",
    "DOCUMENTOS", "FINANCEIRO", "LIMPEZA", "MARKETING",
    "PREVIDENCIARIO", "RECEPÇÃO", "RECURSOS HUMANOS", "REPASSE",
    "TRABALHISTA",
]

MAPA_CENTRO_CUSTO = {
    "0. REPASSE DE CLIENTE": "REPASSE",
    "0. DEVOLUÇÃO DE HONORÁRIOS": "REPASSE",
    "1. REPASSE DE CLIENTE": "REPASSE",
    "0. HONORÁRIOS DE PARCEIROS": "CESSÃO DE CRÉDITO",
    "2. SALÁRIOS": "ADMINISTRATIVO",
    "2. PRÓ-LABORE": "ADMINISTRATIVO",
    "2. VALE TRANSPORTE": "ADMINISTRATIVO",
    "2. VALE ALIMENTAÇÃO": "ADMINISTRATIVO",
    "2. ADIANTAMENTOS": "ADMINISTRATIVO",
    "2. BONIFICAÇÃO": "ADMINISTRATIVO",
    "2. ACORDO DE SÓCIOS": "ADMINISTRATIVO",
    "2. BENEFÍCIOS": "ADMINISTRATIVO",
    "2. SEGURO DE VIDA": "ADMINISTRATIVO",
    "2. SAÚDE OCUPACIONAL": "ADMINISTRATIVO",
    "2. RESCISÃO DO CONTRATO DE TRABALHO": "ADMINISTRATIVO",
    "2. REMUNERAÇÃO DE ADVOGADOS PRESTADORES DE SERVIÇO": "ADMINISTRATIVO",
    "2. REMUNERAÇÃO DE PARCEIROS - PRESTADORES DE SERVIÇO": "ADMINISTRATIVO",
    "3. FGTS": "ADMINISTRATIVO",
    "4. IOF": "ADMINISTRATIVO",
    "4. ISS": "ADMINISTRATIVO",
    "4. INSS": "ADMINISTRATIVO",
    "4. SIMPLES NACIONAL": "ADMINISTRATIVO",
    "5. ALUGUEL": "ADMINISTRATIVO",
    "5. ENERGIA ELÉTRICA": "ADMINISTRATIVO",
    "5. ÁGUA": "ADMINISTRATIVO",
    "5. INTERNET": "ADMINISTRATIVO",
    "5. SOFTWARE E SISTEMAS": "ADMINISTRATIVO",
    "5. CONTADOR": "ADMINISTRATIVO",
    "5. PAGAMENTO DE FATURA": "ADMINISTRATIVO",
    "5. SEGURO DO CARRO": "ADMINISTRATIVO",
    "5. ANUIDADE OAB": "ADMINISTRATIVO",
    "5. ESTACIONAMENTO": "ADMINISTRATIVO",
    "5. SERVIÇOS DE COMODATO": "ADMINISTRATIVO",
    "5. INFORMÁTICA": "ADMINISTRATIVO",
    "6. CONFRATERNIZAÇÕES DO ESCRITÓRIO": "ADMINISTRATIVO",
    "6. CURSOS E TREINAMENTOS": "ADMINISTRATIVO",
    "6. INVESTIMENTO - SALA TATUAPÉ": "ADMINISTRATIVO",
    "6. DIVISÃO DE LUCROS - MENSAL": "ADMINISTRATIVO",
    "6. MATERIAIS DE LIMPEZA E ALIMENTAÇÃO": "ADMINISTRATIVO",
    "6. PASSAGENS AÉREAS E RODOVIÁRIAS": "ADMINISTRATIVO",
    "6. MANUTENÇÃO": "ADMINISTRATIVO",
    "6. GASTOS COM O ESCRITÓRIO": "ADMINISTRATIVO",
    "7. TAXAS BANCÁRIAS": "ADMINISTRATIVO",
    "7. JUROS BANCÁRIOS": "ADMINISTRATIVO",
    "TRANSFERÊNCIA ENTRE CONTAS": "ADMINISTRATIVO",
}

CONTAS = [
    "A - BANCO DO BRASIL - SOCIEDADE",
    "A - CAIXA - ESCRITÓRIO (ESPÉCIE) (SOCIEDADE)",
    "A - ITAU - SOCIEDADE",
    "ASAAS",
    "I - ASAAS - DR. FERREIRA",
    "I - BANCO DO BRASIL - FRANCISCO",
    "I - CAIXA - ESCRITÓRIO (ESPÉCIE) (PF)",
    "I - CAIXA ECONOMICA - DR. FRANCISCO",
    "S - ASAAS - SERVIÇOS",
    "S - ITAU - SERVIÇOS",
]

MAPA_CONTA = {
    "BB": "I - BANCO DO BRASIL - FRANCISCO",
    "BANCO DO BRASIL": "I - BANCO DO BRASIL - FRANCISCO",
    "ITAU": "A - ITAU - SOCIEDADE",
    "ITAÚ": "A - ITAU - SOCIEDADE",
    "CAIXA": "I - CAIXA ECONOMICA - DR. FRANCISCO",
    "ASAAS": "ASAAS",
}


def _resolver_conta(conta_arquivo: str) -> str:
    """Mapeia o nome do arquivo de extrato para o nome da conta no Advbox."""
    nome = conta_arquivo.upper().strip()
    for chave, valor in MAPA_CONTA.items():
        if chave in nome:
            return valor
    return conta_arquivo


def _limpar_uploads_antigos():
    """Remove pastas com mais de 2 horas na pasta de uploads."""
    import shutil
    agora = datetime.now().timestamp()
    for f in os.listdir(UPLOAD_FOLDER):
        caminho = os.path.join(UPLOAD_FOLDER, f)
        if os.path.isdir(caminho) and agora - os.path.getmtime(caminho) > 7200:
            shutil.rmtree(caminho, ignore_errors=True)


def _sanitize(name):
    if ".." in name or "/" in name or "\\" in name:
        return None
    return name


def _get_advbox_client():
    """Cria o AdvboxClient com o token do .env."""
    token = os.environ.get("ADVBOX_TOKEN", "")
    dry_run = os.environ.get("ADVBOX_DRY_RUN", "true").lower() in ("true", "1", "yes")
    return AdvboxClient(token=token, dry_run=dry_run)


# ================================================================
# ROTAS — Páginas
# ================================================================

@app.route("/")
def index():
    advbox_token = os.environ.get("ADVBOX_TOKEN", "")
    return render_template("index.html", advbox_configurado=bool(advbox_token))


@app.route("/conciliar", methods=["POST"])
def conciliar():
    _limpar_uploads_antigos()

    extrato_files = request.files.getlist("extratos")
    advbox_file = request.files.get("advbox")

    if not extrato_files or not extrato_files[0].filename:
        flash("Envie ao menos um extrato bancário (.ofx ou .pdf).", "error")
        return redirect(url_for("index"))

    # --- Salvar arquivos temporários ---
    run_id = uuid.uuid4().hex[:8]
    run_dir = os.path.join(UPLOAD_FOLDER, run_id)
    os.makedirs(run_dir, exist_ok=True)

    extrato_paths = []
    for f in extrato_files:
        if f.filename:
            path = os.path.join(run_dir, f.filename)
            f.save(path)
            extrato_paths.append(path)

    # --- Advbox export é opcional ---
    sistema = []
    erro_advbox = None
    if advbox_file and advbox_file.filename:
        advbox_path = os.path.join(run_dir, advbox_file.filename)
        advbox_file.save(advbox_path)
        try:
            sistema = engine.carregar_advbox_export(advbox_path)
        except Exception as e:
            erro_advbox = f"Não foi possível ler o export do Advbox: {e}. Continuando sem matching."

    extrato = []
    erros = []
    if erro_advbox:
        erros.append(erro_advbox)

    for path in extrato_paths:
        nome = os.path.splitext(os.path.basename(path))[0]
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext == ".ofx":
                movs = engine.carregar_ofx(path, nome)
            elif ext == ".pdf":
                movs = engine.carregar_caixa_pdf(path, nome)
            else:
                erros.append(f"Formato não suportado: {os.path.basename(path)}")
                continue
            extrato.extend(movs)
        except Exception as e:
            erros.append(f"Erro em {os.path.basename(path)}: {e}")

    if not extrato:
        flash("Nenhum movimento encontrado nos extratos enviados.", "error")
        return redirect(url_for("index"))

    # --- Conciliar ---
    baixas, novos, revisao = engine.conciliar(extrato, sistema)

    # --- Montar lista unificada ---
    itens = []
    for i, (mov, l) in enumerate(baixas):
        itens.append({
            "id": f"b-{i}",
            "acao": "baixa",
            "data": mov.data.strftime("%d/%m/%Y"),
            "valor": abs(mov.valor),
            "descricao": mov.descricao,
            "conta": l.get("conta", "") or _resolver_conta(mov.conta),
            "tipo": mov.tipo,
            "categoria": l.get("categoria", ""),
            "centro_custo": l.get("centro_custo", ""),
            "setor": l.get("setor", ""),
            "descricao_advbox": l.get("descricao", ""),
            "pessoa": l.get("pessoa", ""),
            "registro_interno": False,
            "revisar_nota": "",
        })

    for i, l in enumerate(novos):
        cc = l.centro_custo or MAPA_CENTRO_CUSTO.get(l.categoria, "")
        itens.append({
            "id": f"c-{i}",
            "acao": "criar",
            "data": l.vencimento.strftime("%d/%m/%Y"),
            "valor": round(l.valor, 2),
            "descricao": l.descricao,
            "conta": _resolver_conta(l.conta),
            "tipo": l.tipo,
            "categoria": l.categoria,
            "centro_custo": cc,
            "setor": l.setor,
            "descricao_advbox": "",
            "pessoa": l.pessoa,
            "registro_interno": l.registro_interno,
            "revisar_nota": l.revisar,
        })

    for i, (mov, cat) in enumerate(revisao):
        cc_rev = MAPA_CENTRO_CUSTO.get(cat or "", "")
        itens.append({
            "id": f"r-{i}",
            "acao": "revisar",
            "data": mov.data.strftime("%d/%m/%Y"),
            "valor": abs(mov.valor),
            "descricao": mov.descricao,
            "conta": _resolver_conta(mov.conta),
            "tipo": mov.tipo,
            "categoria": cat or "",
            "centro_custo": cc_rev,
            "setor": "",
            "descricao_advbox": "",
            "pessoa": "",
            "registro_interno": False,
            "revisar_nota": "sem regra" if cat is None else "ambíguo",
        })

    # Salvar dados para regenerar Excel depois de edições
    with open(os.path.join(run_dir, "dados.json"), "w", encoding="utf-8") as f:
        json.dump(itens, f, ensure_ascii=False)

    # Gerar Excel inicial
    saida_nome = f"conciliacao_{run_id}.xlsx"
    saida_path = os.path.join(run_dir, saida_nome)
    engine.gerar_planilha(extrato, sistema, saida_path)

    advbox_token = os.environ.get("ADVBOX_TOKEN", "")
    advbox_dry_run = os.environ.get("ADVBOX_DRY_RUN", "true").lower() in ("true", "1", "yes")

    return render_template("resultado.html",
                           itens=itens,
                           categorias=CATEGORIAS,
                           contas=CONTAS,
                           centros_custo=CENTROS_CUSTO,
                           setores=SETORES,
                           total_extrato=len(extrato),
                           total_baixas=len(baixas),
                           total_novos=len(novos),
                           total_revisao=len(revisao),
                           run_id=run_id,
                           erros=erros,
                           advbox_configurado=bool(advbox_token),
                           advbox_dry_run=advbox_dry_run)


# ================================================================
# ROTAS — API (AJAX)
# ================================================================

@app.route("/salvar/<run_id>", methods=["POST"])
def salvar(run_id):
    """Recebe edições do usuário e regenera o Excel."""
    if not _sanitize(run_id):
        return "Inválido", 400

    run_dir = os.path.join(UPLOAD_FOLDER, run_id)
    if not os.path.isdir(run_dir):
        return "Sessão expirada", 404

    itens = request.json
    if not itens:
        return "Sem dados", 400

    # Salvar dados atualizados
    with open(os.path.join(run_dir, "dados.json"), "w", encoding="utf-8") as f:
        json.dump(itens, f, ensure_ascii=False)

    # Regenerar Excel
    _gerar_excel(run_dir, run_id, itens)

    return jsonify({"ok": True, "arquivo": f"conciliacao_{run_id}.xlsx"})


@app.route("/api/advbox/status")
def advbox_status():
    """Verifica se a API do Advbox está configurada e acessível."""
    client = _get_advbox_client()
    if not client.configurado:
        return jsonify({
            "configurado": False,
            "mensagem": "Token do Advbox não configurado. Adicione ADVBOX_TOKEN no arquivo .env",
        })

    try:
        info = client.carregar_settings()
        return jsonify({
            "configurado": True,
            "dry_run": client.dry_run,
            "settings": info,
            "contas": list(client.contas.keys()),
            "categorias": list(client.categorias.keys()),
        })
    except Exception as e:
        return jsonify({
            "configurado": True,
            "dry_run": client.dry_run,
            "erro": str(e),
            "mensagem": "Token configurado mas não foi possível conectar à API.",
        })


@app.route("/api/advbox/enviar/<run_id>", methods=["POST"])
def advbox_enviar(run_id):
    """
    Envia os itens aprovados para o Advbox.
    Espera JSON com a lista de itens (já editados pelo usuário).
    Itens com acao='revisar' são IGNORADOS (nunca postados automaticamente).
    """
    if not _sanitize(run_id):
        return jsonify({"ok": False, "erro": "ID inválido"}), 400

    run_dir = os.path.join(UPLOAD_FOLDER, run_id)
    if not os.path.isdir(run_dir):
        return jsonify({"ok": False, "erro": "Sessão expirada"}), 404

    client = _get_advbox_client()
    if not client.configurado:
        return jsonify({
            "ok": False,
            "erro": "Token do Advbox não configurado. Adicione ADVBOX_TOKEN no .env",
        }), 400

    itens = request.json
    if not itens:
        return jsonify({"ok": False, "erro": "Sem dados"}), 400

    # Carregar settings para mapear nomes → IDs
    try:
        client.carregar_settings()
    except Exception as e:
        return jsonify({
            "ok": False,
            "erro": f"Erro ao carregar settings do Advbox: {e}",
        }), 500

    # Executar conciliação via API
    resultado = client.executar_conciliacao(itens)

    # Salvar log do resultado
    log_path = os.path.join(run_dir, "advbox_resultado.json")
    with open(log_path, "w", encoding="utf-8") as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2, default=str)

    return jsonify({
        "ok": True,
        "dry_run": client.dry_run,
        "resultado": {
            "sucesso": len(resultado["sucesso"]),
            "erros": len(resultado["erros"]),
            "ignorados": len(resultado["ignorados"]),
        },
        "detalhes": resultado,
    })


@app.route("/download/<run_id>/<nome>")
def download(run_id, nome):
    if not _sanitize(run_id) or not _sanitize(nome):
        return "Inválido", 400
    path = os.path.join(UPLOAD_FOLDER, run_id, nome)
    if not os.path.isfile(path):
        return "Arquivo não encontrado", 404
    return send_file(path, as_attachment=True, download_name=nome)


# ================================================================
# HELPERS
# ================================================================

def _gerar_excel(run_dir, run_id, itens):
    """Gera (ou regenera) o Excel a partir da lista de itens editados."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    _AZUL = "1F4E78"
    _HEAD = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    _TXT = Font(name="Arial", size=10)
    _BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)

    def estilo_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(1, c)
            cell.font = _HEAD
            cell.fill = PatternFill("solid", fgColor=_AZUL)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"

    # Resumo
    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = "CONCILIAÇÃO BANCÁRIA"
    ws["A1"].font = Font(name="Arial", bold=True, size=14)
    ws["A3"] = "Gerado em"
    ws["B3"] = datetime.now().strftime("%d/%m/%Y %H:%M")

    baixas = [i for i in itens if i["acao"] == "baixa"]
    criar = [i for i in itens if i["acao"] == "criar"]
    revisar = [i for i in itens if i["acao"] == "revisar"]

    for row, (k, v) in enumerate([
        ("Movimentos no extrato", len(itens)),
        ("Já no Advbox (dar baixa)", len(baixas)),
        ("Criar lançamento", len(criar)),
        ("Revisar", len(revisar)),
    ], start=5):
        ws.cell(row, 1, k).font = _TXT
        ws.cell(row, 2, v).font = Font(name="Arial", bold=True, size=10)

    # Dar Baixa
    ws = wb.create_sheet("DAR BAIXA")
    ws.sheet_properties.tabColor = "548235"
    cols_b = ["Data", "Valor", "Descrição (extrato)", "Conta", "No Advbox", "Cliente"]
    ws.append(cols_b)
    for i in baixas:
        ws.append([i["data"], i["valor"], i["descricao"], i["conta"],
                   i["descricao_advbox"], i["pessoa"]])
    estilo_header(ws, len(cols_b))

    # Criar
    ws = wb.create_sheet("CRIAR LANÇAMENTO")
    ws.sheet_properties.tabColor = _AZUL
    cols_c = ["Conta", "Tipo", "Categoria", "Centro de Custo", "Setor/Unidade",
              "Descrição", "Valor", "Data", "Pessoa", "Registro interno"]
    ws.append(cols_c)
    for i in criar:
        ws.append([i["conta"], i["tipo"], i["categoria"],
                   i.get("centro_custo", ""), i.get("setor", ""),
                   i["descricao"], i["valor"], i["data"], i["pessoa"],
                   "Sim" if i.get("registro_interno") else ""])
    estilo_header(ws, len(cols_c))

    # Revisar
    ws = wb.create_sheet("REVISAR")
    ws.sheet_properties.tabColor = "C00000"
    cols_r = ["Data", "Valor", "Descrição", "Conta", "Tipo", "Sugestão",
              "Centro de Custo", "Setor/Unidade", "Observação"]
    ws.append(cols_r)
    for i in revisar:
        ws.append([i["data"], i["valor"], i["descricao"], i["conta"],
                   i["tipo"], i["categoria"],
                   i.get("centro_custo", ""), i.get("setor", ""),
                   i["revisar_nota"]])
    estilo_header(ws, len(cols_r))

    # Estilo de corpo
    for sh in wb.worksheets:
        if sh.title == "Resumo":
            continue
        for row in sh.iter_rows(min_row=2):
            for cell in row:
                cell.font = _TXT
                cell.border = _BORDER
                cell.alignment = Alignment(vertical="center")
        for col, w in {"A": 30, "B": 14, "C": 32, "D": 32, "E": 14,
                       "F": 14, "G": 26, "H": 14}.items():
            sh.column_dimensions[col].width = w

    saida = os.path.join(run_dir, f"conciliacao_{run_id}.xlsx")
    wb.save(saida)


if __name__ == "__main__":
    print("=== Conciliação Bancária — Interface Web ===")
    print("Acesse: http://localhost:5000\n")
    app.run(debug=True, port=5000)
